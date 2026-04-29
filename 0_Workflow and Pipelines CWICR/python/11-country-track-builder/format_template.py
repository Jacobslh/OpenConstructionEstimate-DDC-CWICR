"""
Extract presentation styling from an existing FORMATTED.xlsx (DE_BERLIN
is the canonical reference) and reapply it to a new dataframe.

Captures: header fill+font, column widths, freeze pane, banded rows.
Cached as JSON so we don't re-parse the reference xlsx on every run.
"""

from __future__ import annotations

import json
from dataclasses import dataclass, field, asdict
from pathlib import Path
from typing import Any

import openpyxl
from openpyxl.styles import Alignment, Border, Fill, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


HERE = Path(__file__).resolve().parent
TEMPLATE_CACHE = HERE / "format_template_cache.json"

REPO_ROOT = HERE.parent.parent.parent
REFERENCE_FORMATTED_XLSX = (
    REPO_ROOT / "DE___DDC_CWICR"
    / "DE_BERLIN_workitems_costs_resources_DDC_CWICR_FORMATTED.xlsx"
)


@dataclass
class FormatTemplate:
    header_fill_rgb: str = "DDDDDD"
    header_font_name: str = "Calibri"
    header_font_size: float = 11
    header_font_bold: bool = True
    column_widths: dict[str, float] = field(default_factory=dict)
    freeze_panes: str = "A2"
    banded_rows_fill_rgb: str | None = "F5F5F5"

    def save(self, path: Path = TEMPLATE_CACHE) -> None:
        path.write_text(json.dumps(asdict(self), indent=2), encoding="utf-8")

    @classmethod
    def load(cls, path: Path = TEMPLATE_CACHE) -> "FormatTemplate":
        if not path.exists():
            return cls()
        data = json.loads(path.read_text(encoding="utf-8"))
        return cls(**data)


def extract_template(xlsx_path: Path = REFERENCE_FORMATTED_XLSX) -> FormatTemplate:
    """Read the reference FORMATTED.xlsx and snapshot its styling."""
    if not xlsx_path.exists():
        return FormatTemplate()

    wb = openpyxl.load_workbook(xlsx_path, read_only=False)
    ws = wb.active

    tpl = FormatTemplate()

    # Header style (row 1).
    header_cell = ws.cell(row=1, column=1)
    if header_cell.fill and header_cell.fill.fgColor:
        rgb = getattr(header_cell.fill.fgColor, "rgb", None)
        if rgb and rgb != "00000000":
            tpl.header_fill_rgb = rgb[-6:]
    if header_cell.font:
        tpl.header_font_name = header_cell.font.name or tpl.header_font_name
        tpl.header_font_size = header_cell.font.size or tpl.header_font_size
        tpl.header_font_bold = bool(header_cell.font.bold)

    # Column widths.
    for col_letter, dim in ws.column_dimensions.items():
        if dim.width is not None:
            tpl.column_widths[col_letter] = float(dim.width)

    # Freeze panes.
    if ws.freeze_panes:
        tpl.freeze_panes = ws.freeze_panes

    return tpl


def ensure_template_cache() -> FormatTemplate:
    """Load the cached template, or extract+save it if missing."""
    if TEMPLATE_CACHE.exists():
        return FormatTemplate.load()
    tpl = extract_template()
    tpl.save()
    return tpl


def apply_template(xlsx_path: Path, tpl: FormatTemplate, n_cols: int) -> None:
    """Open an existing xlsx, apply the template styles, save in place."""
    wb = openpyxl.load_workbook(xlsx_path)
    ws = wb.active

    header_fill = PatternFill(
        start_color=tpl.header_fill_rgb,
        end_color=tpl.header_fill_rgb,
        fill_type="solid",
    )
    header_font = Font(
        name=tpl.header_font_name,
        size=tpl.header_font_size,
        bold=tpl.header_font_bold,
    )
    header_align = Alignment(horizontal="left", vertical="center", wrap_text=True)

    for c in range(1, n_cols + 1):
        cell = ws.cell(row=1, column=c)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_align

    # Apply column widths.
    for col_letter, width in tpl.column_widths.items():
        ws.column_dimensions[col_letter].width = width

    # Default reasonable widths for columns we don't have a template for.
    for c in range(1, n_cols + 1):
        col_letter = get_column_letter(c)
        if col_letter not in tpl.column_widths:
            ws.column_dimensions[col_letter].width = 18

    if tpl.freeze_panes:
        ws.freeze_panes = tpl.freeze_panes

    if tpl.banded_rows_fill_rgb:
        band_fill = PatternFill(
            start_color=tpl.banded_rows_fill_rgb,
            end_color=tpl.banded_rows_fill_rgb,
            fill_type="solid",
        )
        max_row = ws.max_row
        for r in range(2, max_row + 1):
            if r % 2 == 0:
                for c in range(1, n_cols + 1):
                    if not ws.cell(row=r, column=c).fill or \
                            ws.cell(row=r, column=c).fill.fill_type is None:
                        ws.cell(row=r, column=c).fill = band_fill

    wb.save(xlsx_path)
